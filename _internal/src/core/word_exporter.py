"""
Word 表格导出为 Excel

功能：
- 从 Word 文档中提取所有表格
- 导出为单个工作簿（每个表格一个 Sheet）
- 或导出为多个工作簿（每个表格一个文件）
- 保留 Word 表格格式（边框、对齐、字体等）

命名规则：
- 自动检测并跳过公共层级前缀
- 以实际表格名称命名（如"货币资金"而不是"合并财务报表项目注释/货币资金"）
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .word_parser import (
    get_tables_with_paths, NAMESPACES, get_cell_text,
    DEFAULT_SKIP_KEYWORDS
)


def detect_common_prefix(paths):
    """
    检测路径列表的公共前缀层数

    参数:
        paths: 路径列表 ["A/B/C", "A/B/D", ...]

    返回:
        公共前缀层数（0表示没有公共前缀）
    """
    if not paths:
        return 0

    # 将所有路径拆分为层级列表
    all_parts = [path.split('/') for path in paths if path]
    if not all_parts:
        return 0

    min_depth = min(len(parts) for parts in all_parts)

    # 找出公共前缀层数
    common_levels = 0
    for level in range(min_depth):
        level_values = set(parts[level] for parts in all_parts if len(parts) > level)
        if len(level_values) == 1:
            common_levels += 1
        else:
            break

    return common_levels


def trim_path(path, skip_levels):
    """
    去掉路径的前 N 层

    参数:
        path: 原始路径 "A/B/C"
        skip_levels: 要跳过的层数

    返回:
        截取后的路径 "B/C"（如果 skip_levels=1）
    """
    if not path or skip_levels <= 0:
        return path

    parts = path.split('/')
    if skip_levels >= len(parts):
        return parts[-1]  # 至少保留最后一层

    return '/'.join(parts[skip_levels:])


def sanitize_filename(name):
    """
    清理文件名，移除不合法字符

    参数:
        name: 原始名称

    返回:
        合法的文件名
    """
    # Windows 和 macOS 不允许的字符
    invalid_chars = '<>:"/\\|?*'
    result = name
    for char in invalid_chars:
        result = result.replace(char, '_')
    # 移除前后空格
    result = result.strip()
    # 限制长度
    if len(result) > 200:
        result = result[:200]
    return result or 'unnamed'


def sanitize_sheet_name(name):
    """
    清理 Sheet 名称

    Excel Sheet 名称限制：
    - 最多 31 个字符
    - 不能包含 : \\ / ? * [ ] 以及中文冒号
    """
    # 英文非法字符 + 中文冒号
    invalid_chars = ':\\/?*[]：'
    result = name
    for char in invalid_chars:
        result = result.replace(char, '_')
    # 移除前后空格
    result = result.strip()
    # 限制长度
    if len(result) > 31:
        result = result[:31].rstrip()
    return result or 'Sheet'


def get_cell_alignment(cell):
    """
    获取单元格的对齐方式

    返回: (horizontal, vertical)
    """
    horizontal = 'left'
    vertical = 'center'

    # 检查段落对齐
    p = cell.find('.//w:p', NAMESPACES)
    if p is not None:
        pPr = p.find('.//w:pPr', NAMESPACES)
        if pPr is not None:
            jc = pPr.find('.//w:jc', NAMESPACES)
            if jc is not None:
                val = jc.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val', 'left')
                if val == 'center':
                    horizontal = 'center'
                elif val == 'right':
                    horizontal = 'right'
                elif val in ('both', 'distribute'):
                    horizontal = 'justify'

    # 检查单元格垂直对齐
    tcPr = cell.find('.//w:tcPr', NAMESPACES)
    if tcPr is not None:
        vAlign = tcPr.find('.//w:vAlign', NAMESPACES)
        if vAlign is not None:
            val = vAlign.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val', 'center')
            if val == 'top':
                vertical = 'top'
            elif val == 'bottom':
                vertical = 'bottom'
            else:
                vertical = 'center'

    return horizontal, vertical


def get_cell_font(cell):
    """
    获取单元格的字体信息

    返回: dict with 'bold', 'size', 'name'
    """
    font_info = {'bold': False, 'size': 11, 'name': '宋体'}

    # 查找 run 属性
    r = cell.find('.//w:r', NAMESPACES)
    if r is not None:
        rPr = r.find('.//w:rPr', NAMESPACES)
        if rPr is not None:
            # 粗体
            b = rPr.find('.//w:b', NAMESPACES)
            if b is not None:
                val = b.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val', 'true')
                font_info['bold'] = val != 'false' and val != '0'

            # 字号
            sz = rPr.find('.//w:sz', NAMESPACES)
            if sz is not None:
                try:
                    # Word 字号单位是半点，需要除以2
                    size_val = int(sz.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val', '22'))
                    font_info['size'] = size_val // 2
                except ValueError:
                    pass

            # 字体
            rFonts = rPr.find('.//w:rFonts', NAMESPACES)
            if rFonts is not None:
                # 优先使用东亚字体
                font_name = rFonts.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}eastAsia')
                if not font_name:
                    font_name = rFonts.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}ascii')
                if font_name:
                    font_info['name'] = font_name

    return font_info


def get_cell_borders(cell):
    """
    获取单元格的边框信息

    返回: dict with 'top', 'bottom', 'left', 'right'
    """
    borders = {'top': True, 'bottom': True, 'left': True, 'right': True}

    tcPr = cell.find('.//w:tcPr', NAMESPACES)
    if tcPr is not None:
        tcBorders = tcPr.find('.//w:tcBorders', NAMESPACES)
        if tcBorders is not None:
            for side in ['top', 'bottom', 'left', 'right']:
                border_elem = tcBorders.find(f'.//w:{side}', NAMESPACES)
                if border_elem is not None:
                    val = border_elem.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val', 'single')
                    # nil 或 none 表示无边框
                    borders[side] = val not in ('nil', 'none')

    return borders


def get_cell_merge_info(cell):
    """
    获取单元格的合并信息

    返回: (colspan, is_vmerge_start, is_vmerge_continue)
    """
    colspan = 1
    is_vmerge_start = False
    is_vmerge_continue = False

    tcPr = cell.find('.//w:tcPr', NAMESPACES)
    if tcPr is not None:
        # 水平合并（gridSpan）
        gridSpan = tcPr.find('.//w:gridSpan', NAMESPACES)
        if gridSpan is not None:
            try:
                colspan = int(gridSpan.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val', '1'))
            except ValueError:
                pass

        # 垂直合并
        vMerge = tcPr.find('.//w:vMerge', NAMESPACES)
        if vMerge is not None:
            val = vMerge.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val', '')
            if val == 'restart':
                is_vmerge_start = True
            else:
                is_vmerge_continue = True

    return colspan, is_vmerge_start, is_vmerge_continue


def extract_table_data_with_format(table_element):
    """
    从 Word 表格元素提取数据和格式

    参数:
        table_element: Word 表格的 XML 元素

    返回:
        {
            'data': [[cell_info, ...], ...],
            'col_widths': [width1, width2, ...]
        }
        其中 cell_info = {
            'text': str,
            'horizontal': str,
            'vertical': str,
            'bold': bool,
            'font_size': int,
            'font_name': str,
            'borders': dict,
            'colspan': int,
            'is_vmerge_start': bool,
            'is_vmerge_continue': bool
        }
    """
    rows = table_element.findall('.//w:tr', NAMESPACES)
    data = []

    # 获取列宽
    col_widths = []
    tblGrid = table_element.find('.//w:tblGrid', NAMESPACES)
    if tblGrid is not None:
        for gridCol in tblGrid.findall('.//w:gridCol', NAMESPACES):
            width = gridCol.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}w', '0')
            try:
                # Word 宽度单位是 twips (1/20 point)，转换为大约的字符宽度
                twips = int(width)
                char_width = twips / 150  # 大约转换
                col_widths.append(max(8, min(50, char_width)))  # 限制在 8-50 之间
            except ValueError:
                col_widths.append(12)

    for row in rows:
        cells = row.findall('.//w:tc', NAMESPACES)
        row_data = []

        for cell in cells:
            text = get_cell_text(cell)
            h_align, v_align = get_cell_alignment(cell)
            font_info = get_cell_font(cell)
            borders = get_cell_borders(cell)
            colspan, is_vmerge_start, is_vmerge_continue = get_cell_merge_info(cell)

            cell_info = {
                'text': text,
                'horizontal': h_align,
                'vertical': v_align,
                'bold': font_info['bold'],
                'font_size': font_info['size'],
                'font_name': font_info['name'],
                'borders': borders,
                'colspan': colspan,
                'is_vmerge_start': is_vmerge_start,
                'is_vmerge_continue': is_vmerge_continue
            }
            row_data.append(cell_info)

        data.append(row_data)

    return {'data': data, 'col_widths': col_widths}


def extract_table_data(table_element):
    """
    从 Word 表格元素提取数据（简单版本，仅文本）

    参数:
        table_element: Word 表格的 XML 元素

    返回:
        二维列表 [[cell1, cell2, ...], ...]
    """
    rows = table_element.findall('.//w:tr', NAMESPACES)
    data = []

    for row in rows:
        cells = row.findall('.//w:tc', NAMESPACES)
        row_data = []
        for cell in cells:
            text = get_cell_text(cell)
            row_data.append(text)
        data.append(row_data)

    return data


def apply_cell_format(ws, row_idx, col_idx, cell_info):
    """
    将格式应用到 Excel 单元格

    参数:
        ws: Excel worksheet
        row_idx: 行索引（1-based）
        col_idx: 列索引（1-based）
        cell_info: 单元格信息字典
    """
    cell = ws.cell(row=row_idx, column=col_idx)

    # 设置文本
    cell.value = cell_info['text']

    # 设置字体
    cell.font = Font(
        name=cell_info.get('font_name', '宋体'),
        size=cell_info.get('font_size', 11),
        bold=cell_info.get('bold', False)
    )

    # 设置对齐
    cell.alignment = Alignment(
        horizontal=cell_info.get('horizontal', 'left'),
        vertical=cell_info.get('vertical', 'center'),
        wrap_text=True
    )

    # 设置边框
    borders = cell_info.get('borders', {})
    thin = Side(style='thin', color='000000')
    none = Side(style=None)

    cell.border = Border(
        top=thin if borders.get('top', True) else none,
        bottom=thin if borders.get('bottom', True) else none,
        left=thin if borders.get('left', True) else none,
        right=thin if borders.get('right', True) else none
    )


def write_table_to_sheet(ws, table_data):
    """
    将表格数据写入 Excel Sheet，并应用格式

    参数:
        ws: Excel worksheet
        table_data: extract_table_data_with_format 返回的数据
    """
    data = table_data['data']
    col_widths = table_data.get('col_widths', [])

    # 跟踪垂直合并
    vmerge_starts = {}  # (col_idx): start_row

    for row_idx, row_data in enumerate(data, 1):
        col_idx = 1
        for cell_info in row_data:
            # 处理垂直合并的继续单元格（跳过）
            if cell_info.get('is_vmerge_continue', False):
                # 这个单元格是垂直合并的一部分，只需要设置边框
                cell = ws.cell(row=row_idx, column=col_idx)
                borders = cell_info.get('borders', {})
                thin = Side(style='thin', color='000000')
                none = Side(style=None)
                cell.border = Border(
                    top=none,  # 合并单元格中间没有上边框
                    bottom=thin if borders.get('bottom', True) else none,
                    left=thin if borders.get('left', True) else none,
                    right=thin if borders.get('right', True) else none
                )
                col_idx += cell_info.get('colspan', 1)
                continue

            # 应用格式
            apply_cell_format(ws, row_idx, col_idx, cell_info)

            # 处理水平合并
            colspan = cell_info.get('colspan', 1)
            if colspan > 1:
                # 合并单元格
                ws.merge_cells(
                    start_row=row_idx, start_column=col_idx,
                    end_row=row_idx, end_column=col_idx + colspan - 1
                )
                # 为合并后的单元格也设置边框
                for c in range(col_idx + 1, col_idx + colspan):
                    merge_cell = ws.cell(row=row_idx, column=c)
                    borders = cell_info.get('borders', {})
                    thin = Side(style='thin', color='000000')
                    none = Side(style=None)
                    merge_cell.border = Border(
                        top=thin if borders.get('top', True) else none,
                        bottom=thin if borders.get('bottom', True) else none,
                        left=none,
                        right=thin if (c == col_idx + colspan - 1 and borders.get('right', True)) else none
                    )

            # 记录垂直合并起始
            if cell_info.get('is_vmerge_start', False):
                vmerge_starts[col_idx] = row_idx

            col_idx += colspan

    # 处理垂直合并（需要遍历完所有行后才能确定合并范围）
    # 重新遍历一次来处理垂直合并
    vmerge_ranges = {}  # col_idx: (start_row, end_row)
    for row_idx, row_data in enumerate(data, 1):
        col_idx = 1
        for cell_info in row_data:
            if cell_info.get('is_vmerge_start', False):
                vmerge_ranges[col_idx] = [row_idx, row_idx]
            elif cell_info.get('is_vmerge_continue', False):
                if col_idx in vmerge_ranges:
                    vmerge_ranges[col_idx][1] = row_idx
            else:
                # 如果之前有合并，现在结束了
                if col_idx in vmerge_ranges:
                    start_row, end_row = vmerge_ranges[col_idx]
                    if end_row > start_row:
                        ws.merge_cells(
                            start_row=start_row, start_column=col_idx,
                            end_row=end_row, end_column=col_idx
                        )
                    del vmerge_ranges[col_idx]
            col_idx += cell_info.get('colspan', 1)

    # 处理到最后一行仍在合并的单元格
    for col_idx, (start_row, end_row) in vmerge_ranges.items():
        if end_row > start_row:
            ws.merge_cells(
                start_row=start_row, start_column=col_idx,
                end_row=end_row, end_column=col_idx
            )

    # 设置列宽
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # 如果没有从 Word 获取到列宽，自动调整
    if not col_widths and data:
        max_cols = max(sum(c.get('colspan', 1) for c in row) for row in data)
        for idx in range(1, max_cols + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 15


def export_word_tables_to_excel(docx_path, output_path, single_file=True, skip_keywords=None):
    """
    将 Word 文档中的表格导出为 Excel

    参数:
        docx_path: Word 文档路径
        output_path: 输出路径（文件路径或文件夹路径）
        single_file: True=导出为单个工作簿，False=导出为多个工作簿
        skip_keywords: 忽略关键词列表

    返回:
        {
            'success': bool,
            'table_count': int,
            'file_count': int,
            'error': str (如果失败)
        }
    """
    if skip_keywords is None:
        skip_keywords = DEFAULT_SKIP_KEYWORDS

    try:
        # 获取所有表格
        tables = get_tables_with_paths(docx_path, skip_keywords=skip_keywords)

        if not tables:
            return {
                'success': False,
                'table_count': 0,
                'file_count': 0,
                'error': 'Word 文档中没有找到表格'
            }

        # 提取所有路径
        paths = [path for _, path, _ in tables]

        # 检测公共前缀
        common_prefix_levels = detect_common_prefix(paths)

        # 处理后的表格信息（带格式）
        processed_tables = []
        for idx, path, table_elem in tables:
            # 去掉公共前缀
            trimmed_path = trim_path(path, common_prefix_levels)
            # 提取表格数据和格式
            table_data = extract_table_data_with_format(table_elem)
            processed_tables.append({
                'index': idx,
                'original_path': path,
                'name': trimmed_path,
                'table_data': table_data
            })

        if single_file:
            # 导出为单个工作簿
            return _export_single_workbook(processed_tables, output_path)
        else:
            # 导出为多个工作簿
            return _export_multiple_workbooks(processed_tables, output_path)

    except Exception as e:
        return {
            'success': False,
            'table_count': 0,
            'file_count': 0,
            'error': str(e)
        }


def _export_single_workbook(tables, output_path):
    """
    导出为单个工作簿（每个表格一个 Sheet）
    """
    wb = Workbook()
    # 删除默认的 Sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 用于处理重名 Sheet
    sheet_names_used = {}

    for table_info in tables:
        name = table_info['name']
        table_data = table_info['table_data']

        # 生成唯一的 Sheet 名
        base_name = sanitize_sheet_name(name)
        sheet_name = base_name
        if sheet_name in sheet_names_used:
            count = sheet_names_used[sheet_name] + 1
            sheet_names_used[sheet_name] = count
            # 确保加上编号后不超过31字符
            suffix = f"_{count}"
            max_base = 31 - len(suffix)
            sheet_name = base_name[:max_base] + suffix
        else:
            sheet_names_used[sheet_name] = 1

        # 创建 Sheet
        ws = wb.create_sheet(title=sheet_name)

        # 写入数据并应用格式
        write_table_to_sheet(ws, table_data)

    # 保存
    wb.save(output_path)

    return {
        'success': True,
        'table_count': len(tables),
        'file_count': 1,
        'error': None
    }


def _export_multiple_workbooks(tables, output_folder):
    """
    导出为多个工作簿（每个表格一个文件）

    文件命名规则：
    - 如果路径只有一层（如"货币资金"），直接用作文件名
    - 如果路径有多层（如"其他应收款/应收利息"），第一层作为文件名，其余作为 Sheet 名
    """
    # 确保输出文件夹存在
    os.makedirs(output_folder, exist_ok=True)

    # 按第一层分组
    grouped = {}
    for table_info in tables:
        name = table_info['name']
        parts = name.split('/')

        if len(parts) == 1:
            # 单层：文件名 = 表格名，Sheet 名 = 表格名
            file_name = parts[0]
            sheet_name = parts[0]
        else:
            # 多层：第一层作为文件名，其余作为 Sheet 名
            file_name = parts[0]
            sheet_name = '/'.join(parts[1:])

        if file_name not in grouped:
            grouped[file_name] = []
        grouped[file_name].append({
            'sheet_name': sheet_name,
            'table_data': table_info['table_data']
        })

    # 导出每个文件
    file_count = 0
    for file_name, sheets in grouped.items():
        wb = Workbook()
        # 删除默认的 Sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 用于处理重名 Sheet
        sheet_names_used = {}

        for sheet_info in sheets:
            base_name = sanitize_sheet_name(sheet_info['sheet_name'])
            sheet_name = base_name

            if sheet_name in sheet_names_used:
                count = sheet_names_used[sheet_name] + 1
                sheet_names_used[sheet_name] = count
                suffix = f"_{count}"
                max_base = 31 - len(suffix)
                sheet_name = base_name[:max_base] + suffix
            else:
                sheet_names_used[sheet_name] = 1

            # 创建 Sheet
            ws = wb.create_sheet(title=sheet_name)

            # 写入数据并应用格式
            write_table_to_sheet(ws, sheet_info['table_data'])

        # 保存文件
        safe_file_name = sanitize_filename(file_name)
        file_path = os.path.join(output_folder, f"{safe_file_name}.xlsx")
        wb.save(file_path)
        file_count += 1

    return {
        'success': True,
        'table_count': len(tables),
        'file_count': file_count,
        'error': None
    }
