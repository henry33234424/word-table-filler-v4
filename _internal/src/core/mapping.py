"""
映射表生成与管理
- 从 Word 提取表格列表
- 从 Excel 文件夹提取 Sheet 列表
- 自动匹配和手工匹配
- 支持通过 Sheet 标签颜色筛选（绿色标签表示用于填充）
"""

import os
import glob
import pandas as pd
from openpyxl import load_workbook
from .word_parser import get_tables_with_paths, get_table_info


def is_green_tab_color(tab_color):
    """
    判断标签颜色是否为绿色系

    参数:
        tab_color: openpyxl 的 Color 对象

    返回:
        bool: 是否为绿色
    """
    if not tab_color or not tab_color.rgb:
        return False

    rgb = tab_color.rgb
    try:
        # ARGB 格式 (8位) 或 RGB 格式 (6位)
        if len(rgb) == 8:
            r = int(rgb[2:4], 16)
            g = int(rgb[4:6], 16)
            b = int(rgb[6:8], 16)
        elif len(rgb) == 6:
            r = int(rgb[0:2], 16)
            g = int(rgb[2:4], 16)
            b = int(rgb[4:6], 16)
        else:
            return False

        # 绿色判定：G 分量明显大于 R 和 B，且 G 值足够高
        return g > r and g > b and g > 100
    except (ValueError, TypeError):
        return False


def get_sheets_with_color_filter(excel_path):
    """
    获取 Excel 文件的 Sheet 列表，支持颜色筛选

    逻辑：
    - 如果文件中有绿色标签的 Sheet，只返回绿色标签的 Sheet
    - 如果没有绿色标签，返回所有 Sheet

    参数:
        excel_path: Excel 文件路径

    返回:
        (sheet_names, has_green_filter): Sheet 名列表和是否应用了绿色筛选
    """
    try:
        # 注意：必须用 read_only=False 才能读取 sheet_properties
        wb = load_workbook(excel_path, read_only=False, data_only=True)
        all_sheets = []
        green_sheets = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            tab_color = ws.sheet_properties.tabColor
            all_sheets.append(sheet_name)

            if is_green_tab_color(tab_color):
                green_sheets.append(sheet_name)

        wb.close()

        # 如果有绿色标签的 Sheet，只使用绿色的
        if green_sheets:
            return green_sheets, True
        else:
            return all_sheets, False

    except Exception as e:
        print(f"警告: 无法读取标签颜色 {excel_path}: {e}")
        # 回退到 pandas 方式
        try:
            xl = pd.ExcelFile(excel_path)
            return xl.sheet_names, False
        except Exception:
            return [], False


def get_word_tables(docx_path_or_dir, skip_keywords=None):
    """
    获取 Word 模板中所有表格

    参数:
        docx_path_or_dir: .docx 文件路径或已解压的目录路径
        skip_keywords: 要忽略的关键词列表，包含这些关键词的行不会作为表格名称

    返回: [(序号, 路径), ...]
    """
    tables = get_tables_with_paths(docx_path_or_dir, skip_keywords=skip_keywords)
    return [(idx, path) for idx, path, _ in tables]


def get_word_tables_with_info(docx_path_or_dir, skip_keywords=None):
    """
    获取 Word 模板中所有表格及其结构信息

    参数:
        docx_path_or_dir: .docx 文件路径或已解压的目录路径
        skip_keywords: 要忽略的关键词列表，包含这些关键词的行不会作为表格名称

    返回: [(序号, 路径, info), ...]
    info = {'header_rows', 'data_rows', 'columns', 'first_column'}
    """
    tables = get_tables_with_paths(docx_path_or_dir, skip_keywords=skip_keywords)
    result = []
    for idx, path, table_elem in tables:
        info = get_table_info(table_elem)
        result.append((idx, path, info))
    return result


def get_excel_sheets(excel_source, use_color_filter=True):
    """
    获取 Excel 数据源中的所有 Sheet

    参数:
        excel_source: Excel 文件夹路径或单个 Excel 文件路径
        use_color_filter: 是否启用颜色筛选（绿色标签优先）

    返回:
        [(文件名, Sheet名), ...]

    颜色筛选逻辑：
        - 如果 Excel 文件中有绿色标签的 Sheet，只返回绿色标签的
        - 如果没有绿色标签，返回所有 Sheet
    """
    result = []
    filter_stats = {'total_files': 0, 'filtered_files': 0}

    # 判断是单个文件还是文件夹
    if os.path.isfile(excel_source):
        # 单文件模式
        excel_files = [excel_source]
    else:
        # 文件夹模式
        excel_files = glob.glob(os.path.join(excel_source, '*.xlsx'))

    for excel_file in sorted(excel_files):
        file_name = os.path.basename(excel_file)

        # 跳过临时文件
        if file_name.startswith('~$'):
            continue

        filter_stats['total_files'] += 1

        if use_color_filter:
            sheets, was_filtered = get_sheets_with_color_filter(excel_file)
            if was_filtered:
                filter_stats['filtered_files'] += 1
            for sheet_name in sheets:
                result.append((file_name, sheet_name))
        else:
            try:
                xl = pd.ExcelFile(excel_file)
                for sheet_name in xl.sheet_names:
                    result.append((file_name, sheet_name))
            except Exception as e:
                print(f"警告: 无法读取 {file_name}: {e}")

    # 输出筛选统计
    if use_color_filter and filter_stats['filtered_files'] > 0:
        print(f"提示: {filter_stats['filtered_files']}/{filter_stats['total_files']} "
              f"个 Excel 文件使用了绿色标签筛选")

    return result


def get_excel_files(excel_folder):
    """
    获取 Excel 文件夹中的所有文件
    返回: [文件名, ...]
    """
    excel_files = glob.glob(os.path.join(excel_folder, '*.xlsx'))
    result = []
    for excel_file in sorted(excel_files):
        file_name = os.path.basename(excel_file)
        if not file_name.startswith('~$'):
            result.append(file_name)
    return result


def get_sheets_of_file(excel_folder, file_name, use_color_filter=True):
    """
    获取指定 Excel 文件的 Sheet 名列表

    参数:
        excel_folder: Excel 文件夹路径
        file_name: Excel 文件名
        use_color_filter: 是否启用颜色筛选

    返回:
        [Sheet名, ...]
    """
    excel_path = os.path.join(excel_folder, file_name)

    if use_color_filter:
        sheets, _ = get_sheets_with_color_filter(excel_path)
        return sheets
    else:
        try:
            xl = pd.ExcelFile(excel_path)
            return xl.sheet_names
        except Exception:
            return []


def get_all_sheets_of_file(excel_folder, file_name):
    """
    获取指定 Excel 文件的所有 Sheet 名（不筛选）
    用于用户需要查看所有 Sheet 的场景

    返回: [Sheet名, ...]
    """
    excel_path = os.path.join(excel_folder, file_name)
    try:
        xl = pd.ExcelFile(excel_path)
        return xl.sheet_names
    except Exception:
        return []


def preview_excel_data(excel_folder, file_name, sheet_name, max_rows=10):
    """
    预览 Excel 数据
    返回: DataFrame（最多 max_rows 行）
    """
    excel_path = os.path.join(excel_folder, file_name)
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, nrows=max_rows)
        return df
    except Exception as e:
        return None


def normalize_for_match(text):
    """标准化文本用于匹配"""
    if not text:
        return ''
    text = text.replace('-', '/').replace('_', '/').replace('|', '/')
    text = ' '.join(text.split())
    return text.strip()


def apply_word_start_level(word_path, start_level):
    """
    根据起始层级截取 Word 路径

    参数:
        word_path: Word 表格完整路径
        start_level: 起始层级（1-based），1 表示从第一层开始（不截取）

    返回:
        截取后的路径

    示例:
        apply_word_start_level("合并财务报表项目注释/货币资金", 2)
        → "货币资金"

        apply_word_start_level("合并财务报表项目注释/其他应收款/应收利息", 2)
        → "其他应收款/应收利息"
    """
    if not word_path or start_level <= 1:
        return word_path

    parts = word_path.split('/')
    skip = start_level - 1

    if skip >= len(parts):
        return parts[-1]  # 层级超出时返回最后一层

    return '/'.join(parts[skip:])


def remove_prefix_by_separator(text, separator):
    """
    根据分隔符移除前缀

    参数:
        text: 原始文本
        separator: 分隔符

    返回:
        处理后的文本
    """
    if not text or not separator:
        return text

    # 找到第一个分隔符的位置
    pos = text.find(separator)
    if pos != -1:
        # 返回分隔符之后的内容
        return text[pos + len(separator):].strip()

    return text


def try_match(word_path, excel_file, excel_sheet, single_file_mode=False, prefix_separator='', word_start_level=1):
    """
    尝试匹配 Word 路径和 Excel Sheet
    返回: 匹配分数 (0-100)

    匹配策略：
    - 先根据 word_start_level 截取 Word 路径
    - 以 Excel 路径的层数为基准
    - 取 Word 路径的后 N 层（N = Excel 层数）
    - 进行完全匹配

    参数:
        word_path: Word 表格路径
        excel_file: Excel 文件名
        excel_sheet: Sheet 名
        single_file_mode: 是否为单文件模式（单文件模式下只用 Sheet 名匹配）
        prefix_separator: 前缀分隔符（移除分隔符之前的内容）
        word_start_level: Word 路径起始层级（1-based）
    """
    # 应用起始层级截取 Word 路径
    word_path = apply_word_start_level(word_path, word_start_level)
    if single_file_mode:
        # 单文件模式：只用 Sheet 名匹配
        excel_name = excel_sheet
    else:
        # 多文件模式：用文件名（去掉扩展名）
        excel_name = os.path.splitext(excel_file)[0]

    # 应用分隔符规则移除前缀
    excel_name_cleaned = remove_prefix_by_separator(excel_name, prefix_separator)

    # 构建 Excel 路径
    if single_file_mode:
        excel_path = excel_name_cleaned
    else:
        if excel_sheet == excel_name or excel_sheet == excel_name_cleaned:
            excel_path = excel_name_cleaned
        else:
            # Sheet 名也需要应用规则
            sheet_cleaned = remove_prefix_by_separator(excel_sheet, prefix_separator)
            excel_path = f"{excel_name_cleaned}/{sheet_cleaned}"

    word_norm = normalize_for_match(word_path)
    excel_norm = normalize_for_match(excel_path)

    # 完全匹配
    if word_norm == excel_norm:
        return 100

    # 按 Excel 层数截取 Word 路径进行匹配
    excel_parts = excel_norm.split('/')
    word_parts = word_norm.split('/')
    excel_level = len(excel_parts)
    word_level = len(word_parts)

    # 如果 Excel 层数比 Word 多，无法匹配
    if excel_level > word_level:
        return 0

    # 取 Word 的后 N 层（N = Excel 层数）
    word_tail = '/'.join(word_parts[-excel_level:])

    if word_tail == excel_norm:
        return 100

    return 0


def get_level1_title(word_path):
    """从 Word 路径中提取一级标题"""
    if not word_path or word_path == '(无对应Word表格)':
        return ''
    return word_path.split('/')[0]


def get_excel_level1(excel_file, prefix_separator=''):
    """从 Excel 文件名中提取一级标题（应用分隔符规则）"""
    if not excel_file:
        return ''
    name = os.path.splitext(excel_file)[0]
    # 应用分隔符规则
    name = remove_prefix_by_separator(name, prefix_separator)
    # 标准化后取第一层
    name = normalize_for_match(name)
    return name.split('/')[0]


def generate_mapping(word_tables, excel_sheets, existing_mappings=None, match_threshold=100, single_file_mode=False, prefix_separator='', word_start_level=1):
    """
    生成映射表

    参数:
        word_tables: [(序号, 路径), ...]
        excel_sheets: [(文件名, Sheet名), ...]
        existing_mappings: 已有手工映射 {Word路径: (Excel文件名, Sheet名)}
        match_threshold: 自动匹配的最低分数阈值（默认100=完全匹配）
        single_file_mode: 是否为单文件模式（单文件模式下只用 Sheet 名匹配）
        prefix_separator: 前缀分隔符（移除分隔符之前的内容）
        word_start_level: Word 路径起始层级（1-based）

    返回: 映射数据列表 [{序号, Word表格路径, Excel文件名, Sheet名, 匹配分数, 匹配状态}, ...]
    """
    if existing_mappings is None:
        existing_mappings = {}

    groups = {}
    used_excel = set()

    level1_order = []
    for word_idx, word_path in word_tables:
        # 应用起始层级后获取一级标题（用于分组）
        trimmed_path = apply_word_start_level(word_path, word_start_level)
        level1 = get_level1_title(trimmed_path)
        if level1 and level1 not in level1_order:
            level1_order.append(level1)
        if level1 not in groups:
            groups[level1] = {'word_items': [], 'unused_excel': []}

    for word_idx, word_path in word_tables:
        # 应用起始层级
        trimmed_path = apply_word_start_level(word_path, word_start_level)
        level1 = get_level1_title(trimmed_path)

        # 尝试自动匹配
        best_match = None
        best_score = 0

        for excel_file, excel_sheet in excel_sheets:
            key = (excel_file, excel_sheet)
            if key in used_excel:
                continue

            score = try_match(word_path, excel_file, excel_sheet, single_file_mode, prefix_separator, word_start_level)
            if score > best_score:
                best_score = score
                best_match = (excel_file, excel_sheet)

        if best_match and best_score >= match_threshold:
            used_excel.add(best_match)
            groups[level1]['word_items'].append({
                '序号': word_idx,
                'Word表格路径': word_path,
                'Excel文件名': best_match[0],
                'Sheet名': best_match[1],
                '匹配分数': best_score,
                '匹配状态': '已匹配'
            })
        elif word_path in existing_mappings:
            excel_file, sheet_name = existing_mappings[word_path]
            used_excel.add((excel_file, sheet_name))
            groups[level1]['word_items'].append({
                '序号': word_idx,
                'Word表格路径': word_path,
                'Excel文件名': excel_file,
                'Sheet名': sheet_name,
                '匹配分数': 0,
                '匹配状态': '手工匹配'
            })
        else:
            groups[level1]['word_items'].append({
                '序号': word_idx,
                'Word表格路径': word_path,
                'Excel文件名': '',
                'Sheet名': '',
                '匹配分数': 0,
                '匹配状态': '未匹配'
            })

    # 建立 Excel 文件 → Word 一级标题的映射
    excel_to_level1 = {}
    for level1, group_data in groups.items():
        for item in group_data['word_items']:
            excel_file = item.get('Excel文件名', '')
            if excel_file and level1:
                excel_to_level1[excel_file] = level1

    # 将未使用的 Excel sheets 分配到对应的组
    for excel_file, excel_sheet in excel_sheets:
        key = (excel_file, excel_sheet)
        if key not in used_excel:
            if excel_file in excel_to_level1:
                level1 = excel_to_level1[excel_file]
            else:
                level1 = get_excel_level1(excel_file, prefix_separator)

            if level1 not in groups:
                groups[level1] = {'word_items': [], 'unused_excel': []}
                if level1 not in level1_order:
                    level1_order.append(level1)
            groups[level1]['unused_excel'].append({
                '序号': '',
                'Word表格路径': '(无对应Word表格)',
                'Excel文件名': excel_file,
                'Sheet名': excel_sheet,
                '匹配分数': 0,
                '匹配状态': 'Excel未使用'
            })

    # 按一级标题顺序组合结果
    mapping_data = []
    for level1 in level1_order:
        if level1 in groups:
            mapping_data.extend(groups[level1]['word_items'])
            mapping_data.extend(groups[level1]['unused_excel'])

    return mapping_data


def save_mapping_to_excel(mapping_data, output_path):
    """保存映射表到 Excel 文件"""
    df = pd.DataFrame(mapping_data)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='映射表', index=False)
        worksheet = writer.sheets['映射表']
        worksheet.column_dimensions['A'].width = 8
        worksheet.column_dimensions['B'].width = 50
        worksheet.column_dimensions['C'].width = 25
        worksheet.column_dimensions['D'].width = 35
        worksheet.column_dimensions['E'].width = 10
        worksheet.column_dimensions['F'].width = 12


def load_mapping_from_excel(mapping_path):
    """
    从 Excel 文件加载映射表
    返回: {Word路径: (Excel文件名, Sheet名)}
    """
    if not os.path.exists(mapping_path):
        return {}

    try:
        df = pd.read_excel(mapping_path, sheet_name='映射表')
    except Exception:
        return {}

    mappings = {}
    for _, row in df.iterrows():
        word_path = row.get('Word表格路径', '')
        excel_file = row.get('Excel文件名', '')
        sheet_name = row.get('Sheet名', '')

        if (pd.notna(word_path) and str(word_path).strip() and
            word_path != '(无对应Word表格)' and
            pd.notna(excel_file) and str(excel_file).strip() and
            pd.notna(sheet_name) and str(sheet_name).strip()):
            mappings[word_path] = (str(excel_file).strip(), str(sheet_name).strip())

    return mappings


def get_valid_mappings(mapping_data):
    """
    从映射数据中提取有效映射
    返回: [(Word路径, Excel文件名, Sheet名), ...]
    """
    result = []
    for item in mapping_data:
        word_path = item.get('Word表格路径', '')
        excel_file = item.get('Excel文件名', '')
        sheet_name = item.get('Sheet名', '')

        if (word_path and word_path != '(无对应Word表格)' and
            excel_file and sheet_name):
            result.append((word_path, excel_file, sheet_name))

    return result
